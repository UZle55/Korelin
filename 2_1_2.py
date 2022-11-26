import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np


class Input:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        self.do_work()

    def do_work(self):
        dataset = Data(self.file_name, self.vacancy_name)
        statistics1, statistics2, statistics3, statistics4, statistics5, statistics6 = dataset.get_statistics()
        dataset.print_statistics(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)

        report = Report(self.vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        report.generate_excel()
        report.generate_image()


class Vacancy:
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class Data:
    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def get_average_value(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    @staticmethod
    def increase(dictionary, key, amount):
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    @staticmethod
    def print_statistics(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics1))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))

    def get_statistics(self):
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        count_of_vacancies = self.fill_salaries_and_get_count_of_vacancies(salary, salary_of_vacancy_name, salary_city)
        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])
        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])
        statistics1 = self.get_average_value(salary)
        statistics2 = self.get_average_value(salary_of_vacancy_name)
        statistics3 = self.get_average_value(salary_city)
        statistics4 = {}
        for year, salaries in salary_city.items():
            statistics4[year] = round(len(salaries) / count_of_vacancies, 4)
        statistics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])
        return statistics1, vacancies_number, statistics2, vacancies_number_by_name, statistics3, statistics5

    def fill_salaries_and_get_count_of_vacancies(self, salary, salary_of_vacancy_name, salary_city):
        count_of_vacancies = 0
        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increase(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increase(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.increase(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1
        return count_of_vacancies


class Report:
    def __init__(self, vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.statistics1 = statistics1
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.statistics1.keys():
            ws1.append([year, self.statistics1[year], self.statistics3[year], self.statistics2[year], self.statistics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2
        ws2 = self.get_ws2()
        self.set_style(ws1, ws2, data)
        self.wb.save('report.xlsx')

    def get_ws2(self):
        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, column_width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2
        return ws2

    def set_style(self, ws1, ws2, data):
        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.statistics5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.statistics1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def generate_image(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        self.salaries_by_years(ax1)
        self.vacancies_by_years(ax2)
        self.salaries_by_cities(ax3)
        self.vacancies_by_cities(ax4)
        plt.tight_layout()
        plt.savefig('graph.png')

    def salaries_by_years(self, ax1):
        bar1 = ax1.bar(np.array(list(self.statistics1.keys())) - 0.4, self.statistics1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.statistics1.keys())), self.statistics3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.statistics1.keys())) - 0.2, list(self.statistics1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

    def vacancies_by_years(self, ax2):
        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.statistics2.keys())) - 0.4, self.statistics2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.statistics2.keys())), self.statistics4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.statistics2.keys())) - 0.2, list(self.statistics2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

    def salaries_by_cities(self, ax3):
        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.statistics5.keys()))]),
                 list(reversed(list(self.statistics5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

    def vacancies_by_cities(self, ax4):
        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.statistics6.values()])
        ax4.pie(list(self.statistics6.values()) + [other], labels=list(self.statistics6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

Input()
